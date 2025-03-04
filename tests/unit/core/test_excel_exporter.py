"""
エクセル出力エンジンのテスト
"""

import os
import pytest
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.table import Table

from src.core.excel_exporter import ExcelExporter
from src.core.structuring import (
    DocumentStructure,
    CustomerEntry,
    EntryDetail,
    StockInfo,
    QuantityInfo,
)


@pytest.fixture
def sample_data() -> DocumentStructure:
    """テスト用のサンプルデータを提供"""
    return DocumentStructure(
        pdf_filename="test.pdf",
        total_amount="¥10,000",
        customers=[
            CustomerEntry(
                customer_code="C001",
                customer_name="テスト顧客",
                department="営業部",
                box_number="BOX001",
                entries=[
                    EntryDetail(
                        no="001",
                        description="テスト商品1",
                        tax_rate="10%",
                        amount="¥1,000",
                        page_no=1,
                        date_range="2024/01/01-2024/01/31",
                        stock_info=StockInfo(
                            carryover=100,
                            incoming=50,
                            w_value=30,
                            outgoing=20,
                            remaining=130,
                            total=180,
                            unit_price=1000,
                        ),
                        quantity_info=QuantityInfo(quantity=100, unit_price=1000),
                    ),
                    EntryDetail(
                        no="002",
                        description="保管料 文書箱",
                        tax_rate="8%",
                        amount="¥2,000",
                        page_no=1,
                        date_range="2024/02月分(2024/02/15)",
                        stock_info=None,
                        quantity_info=None,
                    ),
                    EntryDetail(
                        no="003",
                        description="荷役料 - 新規入庫 文書箱",
                        tax_rate="10%",
                        amount="¥3,000",
                        page_no=2,
                        date_range=None,
                        stock_info=None,
                        quantity_info=QuantityInfo(quantity=5, unit_price=600),
                    ),
                ],
            )
        ],
    )


class TestExcelExporter:
    """ExcelExporterクラスのテスト"""

    @pytest.fixture
    def exporter(self):
        """ExcelExporterインスタンスを提供"""
        return ExcelExporter()

    def test_style_initialization(self, exporter):
        """スタイル設定の初期化テスト"""
        # ヘッダーの背景色
        assert exporter.header_fill.start_color.rgb == "CCCCCC"
        assert exporter.header_fill.end_color.rgb == "CCCCCC"
        assert exporter.header_fill.fill_type == "solid"

        # フォント
        assert exporter.header_font.bold is True

        # 罫線
        assert all(
            side.style == "thin"
            for side in [
                exporter.border.left,
                exporter.border.right,
                exporter.border.top,
                exporter.border.bottom,
            ]
        )

        # 配置
        assert exporter.center_align.horizontal == "center"

    def test_export_with_filename(self, exporter, sample_data, tmp_path):
        """ファイル名指定でのエクセル出力テスト"""
        output_dir = tmp_path / "excel"
        os.makedirs(output_dir, exist_ok=True)
        output_path = str(output_dir / "test_output.xlsx")

        # エクセルファイルの出力
        exporter.export(sample_data, output_path)

        # ファイルの存在確認
        assert os.path.exists(output_path)
        assert os.path.basename(output_path) == "test_output.xlsx"

        # ワークブックの検証
        wb = load_workbook(output_path)
        ws = wb.active

        # シート名の確認
        assert ws.title == "請求書データ"

        # ヘッダー行の検証
        headers = [cell.value for cell in ws[1]]
        assert "PDF名" in headers
        assert "取引先コード" in headers
        assert "明細番号" in headers
        assert "在庫情報_繰越" in headers
        assert "数量情報_数量" in headers
        assert "業務タイプ" in headers
        assert "業務詳細" in headers
        assert "業務日付" in headers

        # データ行の検証（1行目）
        assert ws.cell(row=2, column=1).value == "test.pdf"  # PDF名
        assert ws.cell(row=2, column=3).value == "C001"  # 顧客コード
        assert ws.cell(row=2, column=7).value == "001"  # 明細番号
        assert ws.cell(row=2, column=13).value == 100  # 繰越在庫
        assert ws.cell(row=2, column=20).value == 100  # 数量

        # データ行の検証（2行目 - 保管料）
        assert ws.cell(row=3, column=8).value == "保管料 文書箱"  # 商品名
        assert ws.cell(row=3, column=22).value == "保管"  # 業務タイプ
        assert ws.cell(row=3, column=23).value == "保管"  # 業務詳細
        assert ws.cell(row=3, column=24).value == "2024/02/15"  # 業務日付

        # データ行の検証（3行目 - 荷役料）
        assert ws.cell(row=4, column=8).value == "荷役料 - 新規入庫 文書箱"  # 商品名
        assert ws.cell(row=4, column=22).value == "荷役"  # 業務タイプ
        assert ws.cell(row=4, column=23).value == "新規入庫"  # 業務詳細

        # スタイルの検証
        # ヘッダー行
        for cell in ws[1]:
            assert cell.fill.start_color.rgb == "CCCCCC"
            assert cell.font.bold is True
            assert cell.border.left.style == "thin"
            assert cell.alignment.horizontal == "center"

        # データ行
        for row in ws.iter_rows(min_row=2, max_row=4):
            for cell in row:
                assert cell.border.left.style == "thin"

        # テーブルの検証
        assert len(ws.tables) == 1
        table = list(ws.tables.values())[0]
        assert table.name == "InvoiceTable"
        assert table.ref == f"A1:{get_column_letter(len(exporter.headers))}{ws.max_row}"

        # フィルタヘルパーシートの検証
        assert "フィルタヘルパー" in wb.sheetnames
        helper_sheet = wb["フィルタヘルパー"]
        assert helper_sheet["A1"].value == "フィルタと名前付き範囲の使用方法"
        assert helper_sheet["A4"].value == "InvoiceTable - 全データを含むテーブル"

        # 名前付き範囲の検証
        assert "AllData" in wb.defined_names
        assert "StorageData" in wb.defined_names
        assert "HandlingData" in wb.defined_names
        assert "TransportData" in wb.defined_names

    def test_export_without_filename(self, exporter, sample_data, tmp_path):
        """ファイル名未指定でのエクセル出力テスト"""
        output_dir = tmp_path / "excel"
        os.makedirs(output_dir, exist_ok=True)
        output_path = str(output_dir / "auto_generated.xlsx")

        # エクセルファイルの出力
        exporter.export(sample_data, output_path)

        # ファイルの存在確認
        assert os.path.exists(output_path)
        assert output_path.endswith(".xlsx")

    def test_export_with_invalid_path(self, exporter, sample_data):
        """無効なパスでのエクセル出力テスト"""
        with pytest.raises(OSError):
            # Windowsで無効なパス文字を含むパスを指定
            exporter.export(sample_data, "invalid/path/*:<>")

    def test_export_with_empty_data(self, exporter, tmp_path):
        """空のデータでのエクセル出力テスト"""
        output_dir = tmp_path / "excel"
        os.makedirs(output_dir, exist_ok=True)
        output_path = str(output_dir / "empty_data.xlsx")
        
        empty_data = DocumentStructure(
            pdf_filename="test.pdf",
            total_amount="¥0",
            customers=[],
        )

        # エクセルファイルの出力
        exporter.export(empty_data, output_path)

        # ファイルの存在確認
        assert os.path.exists(output_path)

        # ワークブックの検証
        wb = load_workbook(output_path)
        ws = wb.active

        # ヘッダー行のみ存在することを確認
        assert ws.max_row == 1
        
        # テーブルが作成されていないことを確認（データがないため）
        assert len(ws.tables) == 0
        
        # フィルタヘルパーシートは存在する
        assert "フィルタヘルパー" in wb.sheetnames

    def test_extract_date_from_period(self, exporter):
        """期間から日付を抽出するメソッドのテスト"""
        # 通常のケース
        assert exporter._extract_date_from_period("2024/03月分(2024/03/15)") == "2024/03/15"
        
        # 範囲がある場合
        assert exporter._extract_date_from_period("2024/04月分(2024/04/01 - 2024/04/30)") == "2024/04/01"
        
        # 括弧がない場合
        assert exporter._extract_date_from_period("2024/05月分") == ""
        
        # Noneの場合
        assert exporter._extract_date_from_period(None) == ""
        
        # 空文字の場合
        assert exporter._extract_date_from_period("") == ""

    def test_get_product_type_info(self, exporter):
        """商品名から業務タイプと詳細を取得するメソッドのテスト"""
        # 保管料のケース
        result = exporter._get_product_type_info("保管料 文書箱")
        assert result["type"] == "保管"
        assert result["detail"] == "保管"
        
        # 荷役料のケース
        result = exporter._get_product_type_info("荷役料 - 新規入庫 文書箱")
        assert result["type"] == "荷役"
        assert result["detail"] == "新規入庫"
        
        # 運搬料のケース
        result = exporter._get_product_type_info("運搬料 - 寺田便 文書箱")
        assert result["type"] == "運搬"
        assert result["detail"] == "入出庫"
        
        # マッピングにない商品名
        result = exporter._get_product_type_info("未知の商品")
        assert result["type"] == ""
        assert result["detail"] == ""
        
        # Noneの場合
        result = exporter._get_product_type_info(None)
        assert result["type"] == ""
        assert result["detail"] == ""
