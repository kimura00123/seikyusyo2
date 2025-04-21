import os
import re  # reモジュールをインポート
from typing import Dict, Any, Union, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from .structuring import DocumentStructure, CustomerEntry, EntryDetail
# mappings モジュールから定義をインポート
from .mappings import product_type_mapping, regex_patterns
from src.utils.config import get_settings
import logging # logging をインポート (任意)

settings = get_settings()
logger = logging.getLogger(__name__) # logger を設定 (任意)

class ExcelExporter:
    def __init__(self):
        self.workbook = Workbook()
        # デフォルトのシートを取得または作成
        if "Sheet" in self.workbook.sheetnames:
            self.sheet = self.workbook["Sheet"]
            self.sheet.title = "請求データ"
        else:
            self.sheet = self.workbook.active
            self.sheet.title = "請求データ"

        self._setup_styles()

        # ヘッダー情報を定義 (フィルタ補助列を含む)
        self.headers = [
            "PDFファイル名", "合計金額", "取引先コード", "取引先名", "部署", "箱番",
            "No", "摘要", "税率", "金額", "期間", "ページ番号",
            "繰越", "入庫", "W", "出庫", "残", "合計", "単価(在庫)",
            "数量", "単価(数量)",
            "業務区分", "業務内容", "業務日付" # フィルタ補助列
        ]

    def _setup_styles(self):
        """スタイルの設定"""
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        self.center_align = Alignment(horizontal="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _extract_date_from_period(self, period: str) -> str:
        """期間から日付を抽出する"""
        if not period:
            return ""
            
        # 括弧内の日付を抽出 例: 2024/08月分(2024/08/09) -> 2024/08/09
        try:
            if "(" in period and ")" in period:
                date_part = period.split("(")[1].split(")")[0]
                # 範囲がある場合は最初の日付を使用
                if " - " in date_part:
                    date_part = date_part.split(" - ")[0]
                # 日付形式か確認 (簡易チェック)
                if re.match(r'\d{4}/\d{2}/\d{2}', date_part):
                    return date_part
            return ""
        except Exception as e:
            logger.warning(f"期間からの日付抽出エラー: {period}, エラー: {e}")
            return ""

    def _get_product_type_info(self, product_name: Optional[str]) -> Dict[str, str]:
        """
        商品名から業務区分と業務内容を取得する (共通化された定義を使用)
        """
        default_result = {"type": "", "detail": ""}
        if not product_name:
            return default_result

        # product_type_mapping で最も長く一致するキーを探す
        best_match_key = None
        for key in product_type_mapping:
            if key in product_name:
                if best_match_key is None or len(key) > len(best_match_key):
                    best_match_key = key

        if best_match_key:
            return product_type_mapping[best_match_key].copy() # コピーを返す

        # マッピングで見つからない場合、正規表現パターンを試す
        for pattern, info in regex_patterns:
            if pattern.search(product_name):
                return info.copy() # コピーを返す

        # どのパターンにもマッチしない場合
        # logger.warning(f"商品名「{product_name}」に対応する業務区分/内容が見つかりません。") # 必要に応じてログ出力
        return default_result

    def export(
        self, document: Union[DocumentStructure, Dict[str, Any]], output_path: str
    ) -> None:
        """エクセルファイルを出力する"""
        # ヘッダー行の書き込み
        for col, header in enumerate(self.headers, 1):
            cell = self.sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # データの書き込み
        row = 2
        # documentがdict型の場合は直接アクセス、そうでない場合は属性アクセス
        doc_data = document if isinstance(document, dict) else document.model_dump()

        for customer in doc_data.get("customers", []):
            for entry in customer.get("entries", []):
                # 基本情報の書き込み (列番号を修正)
                self.sheet.cell(row=row, column=1).value = doc_data.get("pdf_filename")
                self.sheet.cell(row=row, column=2).value = doc_data.get("total_amount")
                self.sheet.cell(row=row, column=3).value = customer.get("customer_code")
                self.sheet.cell(row=row, column=4).value = customer.get("customer_name")
                self.sheet.cell(row=row, column=5).value = customer.get("department")
                self.sheet.cell(row=row, column=6).value = customer.get("box_number")
                self.sheet.cell(row=row, column=7).value = entry.get("no")
                self.sheet.cell(row=row, column=8).value = entry.get("description")
                self.sheet.cell(row=row, column=9).value = entry.get("tax_rate")
                self.sheet.cell(row=row, column=10).value = entry.get("amount")
                self.sheet.cell(row=row, column=11).value = entry.get("date_range")
                self.sheet.cell(row=row, column=12).value = entry.get("page_no")

                # 在庫情報の書き込み (列番号を修正)
                stock_info = entry.get("stock_info")
                if stock_info:
                    self.sheet.cell(row=row, column=13).value = stock_info.get("carryover")
                    self.sheet.cell(row=row, column=14).value = stock_info.get("incoming")
                    self.sheet.cell(row=row, column=15).value = stock_info.get("w_value")
                    self.sheet.cell(row=row, column=16).value = stock_info.get("outgoing")
                    self.sheet.cell(row=row, column=17).value = stock_info.get("remaining")
                    self.sheet.cell(row=row, column=18).value = stock_info.get("total")
                    self.sheet.cell(row=row, column=19).value = stock_info.get("unit_price")

                # 数量情報の書き込み (列番号を修正)
                quantity_info = entry.get("quantity_info")
                if quantity_info:
                    self.sheet.cell(row=row, column=20).value = quantity_info.get("quantity")
                    self.sheet.cell(row=row, column=21).value = quantity_info.get("unit_price")

                # フィルタ補助用の情報を書き込み (列番号を修正)
                product_type_info = self._get_product_type_info(entry.get("description", ""))
                self.sheet.cell(row=row, column=22).value = product_type_info["type"]
                self.sheet.cell(row=row, column=23).value = product_type_info["detail"]
                self.sheet.cell(row=row, column=24).value = self._extract_date_from_period(entry.get("date_range", ""))

                # セルにボーダーを適用
                for col_idx in range(1, len(self.headers) + 1):
                    self.sheet.cell(row=row, column=col_idx).border = self.border

                row += 1

        # オートフィルタの適用
        if row > 1: # データ行が存在する場合のみフィルタを設定
             self.sheet.auto_filter.ref = f"A1:{get_column_letter(len(self.headers))}{row-1}"
        else:
             # データがない場合はフィルタ参照を解除または設定しない
             if self.sheet.auto_filter.ref:
                 self.sheet.auto_filter.ref = None

        # 名前付き範囲を定義
        self._create_named_ranges()

        # フィルタ用補助シートを追加
        self._add_helper_sheet()

        # 列幅の自動調整
        for col_idx in range(1, len(self.headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            # ヘッダーの長さを初期値とする
            header_value = self.headers[col_idx-1]
            if header_value:
                 max_length = len(str(header_value))

            # データ行の長さを確認
            # iter_rows を使う方が効率的
            for r in range(2, row): # row は次の書き込み行なので row-1 まで
                 cell_value = self.sheet.cell(row=r, column=col_idx).value
                 if cell_value:
                     cell_len = len(str(cell_value))
                     if cell_len > max_length:
                         max_length = cell_len

            # 少し余裕を持たせる
            adjusted_width = max_length + 2
            # 最大幅を設定 (任意)
            # adjusted_width = min(adjusted_width, 50)
            self.sheet.column_dimensions[column_letter].width = adjusted_width

        # ファイルの保存
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.workbook.save(output_path)

    def _create_named_ranges(self):
        """基本的な名前付き範囲を定義"""
        max_row = self.sheet.max_row
        if max_row < 1: # シートが空の場合は何もしない
             return

        # 全体のデータ範囲 (ヘッダーを含む)
        all_data_ref = f"$A$1:${get_column_letter(len(self.headers))}${max_row}"
        self.workbook.create_named_range("AllData", self.sheet, all_data_ref)

        # 他の名前付き範囲は削除またはコメントアウト (FILTER関数を使わないため)
        # self.workbook.create_named_range("StorageData", ...)
        # self.workbook.create_named_range("HandlingData", ...)
        # self.workbook.create_named_range("TransportData", ...)

        # 既存の同名範囲があれば上書きされる

    def _add_helper_sheet(self):
        """フィルタ補助シートを追加"""
        # 既存のシートがあれば削除して再作成 (任意)
        if "フィルタヘルパー" in self.workbook.sheetnames:
            del self.workbook["フィルタヘルパー"]
        helper_sheet = self.workbook.create_sheet(title="フィルタヘルパー")

        # 説明テキスト (内容は現状維持)
        helper_sheet["A1"] = "フィルタと名前付き範囲の使用方法"
        helper_sheet["A1"].font = Font(bold=True, size=14)
        helper_sheet["A3"] = "主な名前付き範囲："
        helper_sheet["A3"].font = Font(bold=True)
        helper_sheet["A4"] = "AllData - 全データを含む範囲 (ヘッダー含む)"
        # helper_sheet["A5"] = "StorageData - 保管業務のみのデータ" # コメントアウト
        # helper_sheet["A6"] = "HandlingData - 荷役業務のみのデータ" # コメントアウト
        # helper_sheet["A7"] = "TransportData - 運搬業務のみのデータ" # コメントアウト

        helper_sheet["A9"] = "フィルタの使用方法："
        helper_sheet["A9"].font = Font(bold=True)
        helper_sheet["A10"] = "1. データシートのヘッダー行にある▼ボタンでフィルタリングできます。"
        helper_sheet["A11"] = "2. 「業務区分」や「業務内容」列で特定の業務を絞り込めます。"

        helper_sheet["A13"] = "補助列について："
        helper_sheet["A13"].font = Font(bold=True)
        helper_sheet["A14"] = "業務区分: 保管、荷役、運搬などの大分類"
        helper_sheet["A15"] = "業務内容: 新規入庫、入庫、出庫、廃棄などの詳細分類"
        helper_sheet["A16"] = "業務日付: 期間から抽出した実際の日付（日付フィルタ用）"

        # 列幅の調整
        helper_sheet.column_dimensions["A"].width = 80 # 少し短く調整
